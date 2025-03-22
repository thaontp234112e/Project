import os
import sys
from PyQt6.QtWidgets import (QDialog, QApplication, QPushButton, QTableWidgetItem,
                             QHeaderView, QFileDialog, QMessageBox)
from PyQt6.QtGui import QColor, QBrush
from PyQt6.uic import loadUi
import sys
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from models.attendance_manager import QuanLyDiemDanh
from statistic_windowExt import Ui_StatisticDialog
from models.admin_accounts import AdminAccount

class Ui_ManageDialog(QDialog):
    def __init__(self):
        super(Ui_ManageDialog, self).__init__()
        current_dir = os.path.dirname(os.path.abspath(__file__))
        ui_file = os.path.join(current_dir, "manage_window.ui")
        loadUi(ui_file, self)
        self.quan_ly_diem_danh = QuanLyDiemDanh()
        self.quan_ly_diem_danh.doc_tu_csv('../dataset/Attendance.csv')
        self.pushButtonClose.clicked.connect(self.close)
        self.pushButtonImport.clicked.connect(self.import_excel)
        self.pushButtonStatistic.clicked.connect(self.open_statistic)
        self.pushButtonShowAll.clicked.connect(self.hien_thi_tat_ca_sinh_vien)
        self.thiet_lap_bang()
        self.hien_thi_danh_sach_ngay()
        self.selected_date = None
        if self.quan_ly_diem_danh.lay_danh_sach_ngay():
            self.selected_date = self.quan_ly_diem_danh.lay_danh_sach_ngay()[0]

    def thiet_lap_bang(self):
        self.tableWidget.setColumnCount(5)
        self.tableWidget.setHorizontalHeaderLabels(["Student ID", "Name", "Time", "Status", "Image registed"])
        header = self.tableWidget.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)

        # Make sure the header is visible
        header.setVisible(True)

        self.tableWidget.clicked.connect(self.handle_student_click)

    def hien_thi_danh_sach_ngay(self):

        for i in reversed(range(self.dateButtonLayout.count())):
            widget = self.dateButtonLayout.itemAt(i).widget()
            if widget is not None:
                widget.deleteLater()
        danh_sach_ngay = self.quan_ly_diem_danh.lay_danh_sach_ngay()
        danh_sach_ngay.sort(reverse=True)

        for ngay in danh_sach_ngay:
            button = QPushButton(ngay)
            button.setMinimumHeight(30)
            # Add stylesheet to the date buttons
            button.setStyleSheet("""
                QPushButton {
                    background-color: #E0EEFF;
                    color: #0E1743;
                    font: 57 9pt "Montserrat Medium";
                    border-radius: 5px;
                    padding: 5px;
                    text-align: center;
                    padding-left: 10px;
                }
                QPushButton:hover {
                    background-color: #E0EEFF;
                }
                QPushButton:pressed {
                    background-color: #B0D4FF;
                }
            """)
            button.clicked.connect(lambda checked, n=ngay: self.hien_thi_diem_danh_theo_ngay(n))
            self.dateButtonLayout.addWidget(button)
    def hien_thi_tat_ca_sinh_vien(self):
        self.label_selected_date.setText("Student List")

        self.tableWidget.setColumnCount(3)
        self.tableWidget.setHorizontalHeaderLabels(["Student ID", "Name", "Image registed"])

        header = self.tableWidget.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)

        danh_sach_sv = self.quan_ly_diem_danh.lay_danh_sach_sinh_vien()

        self.tableWidget.setRowCount(len(danh_sach_sv))

        for row, sv in enumerate(danh_sach_sv):
            ma_sv = sv.ma_sv
            ho_ten = sv.ho_ten

            self.tableWidget.setItem(row, 0, QTableWidgetItem(ma_sv))
            self.tableWidget.setItem(row, 1, QTableWidgetItem(ho_ten))

            da_dang_ky_anh = sv.da_dang_ky_anh()
            item_dang_ky = QTableWidgetItem("Yes" if da_dang_ky_anh else "No")
            self.tableWidget.setItem(row, 2, item_dang_ky)

            if not da_dang_ky_anh:
                for col in range(3):
                    cell_item = self.tableWidget.item(row, col)
                    cell_item.setBackground(QBrush(QColor(215,255,247)))

        self.tableWidget.itemDoubleClicked.connect(self.open_student_detail)

    def hien_thi_diem_danh_theo_ngay(self, ngay):
        self.selected_date = ngay
        self.label_selected_date.setText(f"Date: {ngay}")

        # Set the correct number of columns and proper headers
        self.tableWidget.setColumnCount(6)
        self.tableWidget.setHorizontalHeaderLabels([
            "Student ID", "Name",
            "Check-in", "Status In",
            "Check-out", "Status Out"
        ])

        # Make sure the header is properly formatted
        header = self.tableWidget.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(5, QHeaderView.ResizeMode.ResizeToContents)


        # Make sure the header is visible
        header.setVisible(True)

        student_names = {}
        if os.path.exists('../dataset/students.csv'):
            try:
                with open('../dataset/students.csv', newline='', encoding='utf-8') as f:
                    import csv
                    reader = csv.reader(f)
                    for row in reader:
                        if len(row) >= 2:
                            student_names[row[0]] = row[1]
            except Exception as e:
                print(f"Error reading students.csv file: {e}")

        danh_sach_diem_danh = self.quan_ly_diem_danh.loc_theo_ngay(ngay)

        sv_dict = {}
        for item in danh_sach_diem_danh:
            ma_sv = item['ma_sv']
            if ma_sv not in sv_dict:
                sv_dict[ma_sv] = {'check_in': None, 'check_out': None}

            trang_thai = item.get('trang_thai', '').lower()
            if 'in' in trang_thai:
                sv_dict[ma_sv]['check_in'] = item
            elif 'out' in trang_thai:
                sv_dict[ma_sv]['check_out'] = item
            else:
                if sv_dict[ma_sv]['check_in'] is None:
                    sv_dict[ma_sv]['check_in'] = item
                else:
                    sv_dict[ma_sv]['check_out'] = item

        for ma_sv, ho_ten in student_names.items():
            if ma_sv not in sv_dict:
                sv_dict[ma_sv] = {'check_in': None, 'check_out': None}

        self.tableWidget.setRowCount(len(sv_dict))

        row = 0
        for ma_sv, data in sv_dict.items():
            check_in_data = data['check_in']
            check_out_data = data['check_out']

            self.tableWidget.setItem(row, 0, QTableWidgetItem(ma_sv))

            ho_ten = student_names.get(ma_sv, ma_sv)
            self.tableWidget.setItem(row, 1, QTableWidgetItem(ho_ten))

            color_absent = QColor(255, 200, 200)

            if check_in_data:
                thoi_gian_parts = check_in_data.get('thoi_gian', '').split() if check_in_data.get('thoi_gian') else ["",
                                                                                                                     ""]
                if len(thoi_gian_parts) > 1:
                    thoi_gian_hien_thi = thoi_gian_parts[1]
                else:
                    thoi_gian_hien_thi = check_in_data.get('thoi_gian', '')

                item_checkin_time = QTableWidgetItem(thoi_gian_hien_thi)
                self.tableWidget.setItem(row, 2, item_checkin_time)

                trang_thai_in = check_in_data.get('trang_thai', 'Clock In')
                item_checkin_status = QTableWidgetItem(trang_thai_in)
                self.tableWidget.setItem(row, 3, item_checkin_status)
            else:
                item_checkin_time = QTableWidgetItem("Not check-in")
                item_checkin_status = QTableWidgetItem("Absence")
                self.tableWidget.setItem(row, 2, item_checkin_time)
                self.tableWidget.setItem(row, 3, item_checkin_status)
                item_checkin_time.setBackground(QBrush(color_absent))
                item_checkin_status.setBackground(QBrush(color_absent))

            if check_out_data:
                thoi_gian_parts = check_out_data.get('thoi_gian', '').split() if check_out_data.get('thoi_gian') else [
                    "", ""]
                if len(thoi_gian_parts) > 1:
                    thoi_gian_hien_thi = thoi_gian_parts[1]
                else:
                    thoi_gian_hien_thi = check_out_data.get('thoi_gian', '')

                item_checkout_time = QTableWidgetItem(thoi_gian_hien_thi)
                self.tableWidget.setItem(row, 4, item_checkout_time)

                trang_thai_out = check_out_data.get('trang_thai', 'Clock Out')
                item_checkout_status = QTableWidgetItem(trang_thai_out)
                self.tableWidget.setItem(row, 5, item_checkout_status)
            else:
                item_checkout_time = QTableWidgetItem("Not check-out")
                item_checkout_status = QTableWidgetItem("Absence")
                self.tableWidget.setItem(row, 4, item_checkout_time)
                self.tableWidget.setItem(row, 5, item_checkout_status)
                item_checkout_time.setBackground(QBrush(color_absent))
                item_checkout_status.setBackground(QBrush(color_absent))

            row += 1

    def import_excel(self):
        file_dialog = QFileDialog()
        file_path, _ = file_dialog.getOpenFileName(
            self, "Select Excel file", "", "Excel Files (*.xlsx *.xls)"
        )
        if file_path:
            try:
                from openpyxl import load_workbook

                wb = load_workbook(filename=file_path, read_only=True)
                ws = wb.active

                header_row = next(ws.iter_rows(values_only=True))
                header_row = [str(cell).lower() if cell else "" for cell in header_row]

                student_id_col = None
                fullname_col = None

                for idx, header in enumerate(header_row):
                    if header in ["studentid", "student_id", "masv", "mã sv", "ma_sv", "id"]:
                        student_id_col = idx
                    elif header in ["fullname", "full_name", "hoten", "họ tên", "ho_ten", "name"]:
                        fullname_col = idx

                if student_id_col is None or fullname_col is None:
                    QMessageBox.warning(self, "Error", "StudentID or FullName column not found in the Excel file.")
                    return False

                student_list = []
                for row in list(ws.iter_rows(values_only=True))[1:]:
                    if row[student_id_col] and row[fullname_col]:
                        student_id = str(row[student_id_col]).strip()
                        full_name = str(row[fullname_col]).strip()
                        student_list.append([student_id, full_name])

                if student_list:
                    with open('../dataset/students.csv', 'w', newline='', encoding='utf-8') as f:
                        import csv
                        writer = csv.writer(f)
                        for student in student_list:
                            writer.writerow(student)

                    self.quan_ly_diem_danh.load_student_info()

                    QMessageBox.information(self, "Success", f"Imported {len(student_list)} students from the Excel file.")
                    self.hien_thi_tat_ca_sinh_vien()
                    return True
                else:
                    QMessageBox.warning(self, "Warning", "No student data found in the Excel file.")
                    return False

            except Exception as e:
                QMessageBox.warning(self,"Error", f"Unable to process the Excel file: {str(e)}")
                return False
    def open_statistic(self):
        statistic_dialog = Ui_StatisticDialog(self.quan_ly_diem_danh)
        statistic_dialog.exec()

    def handle_student_click(self, index):
        row = index.row()

        student_id = self.tableWidget.item(row, 0).text()
        student_name = self.tableWidget.item(row, 1).text()

        from student_detail_windowExt import Ui_StudentDetailDialogExt
        detail_dialog = Ui_StudentDetailDialogExt(student_id, student_name, self.quan_ly_diem_danh)
        detail_dialog.exec()

        if self.selected_date:
            self.hien_thi_diem_danh_theo_ngay(self.selected_date)
        else:
            self.hien_thi_tat_ca_sinh_vien()
    def open_student_detail(self, item):
        row = item.row()

        student_id = self.tableWidget.item(row, 0).text()
        student_name = self.tableWidget.item(row, 1).text()

        from student_detail_windowExt import Ui_StudentDetailDialogExt
        detail_dialog = Ui_StudentDetailDialogExt(student_id, student_name, self.quan_ly_diem_danh)
        detail_dialog.exec()

        if self.selected_date:
            self.hien_thi_diem_danh_theo_ngay(self.selected_date)
        else:
            self.hien_thi_tat_ca_sinh_vien()

def main():
    app = QApplication(sys.argv)
    window = Ui_ManageDialog()
    window.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()