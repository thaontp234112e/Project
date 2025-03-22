import sys
import os
from PyQt6.QtWidgets import (QApplication, QTableWidgetItem,
                             QHeaderView, QFileDialog, QMessageBox,
                             QDialog, QVBoxLayout, QTableWidget, QLabel, QPushButton)
from PyQt6.QtGui import QColor, QBrush
from statistic_window import Ui_StatisticDialog
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
import csv

# Thêm đường dẫn đến thư mục cha để import module từ thư mục models
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))


class StudentAttendanceDetailDialog(QDialog):
    """Dialog to display detailed student attendance for a specific date"""

    def __init__(self, date, attendance_records, student_list, parent=None):
        super(StudentAttendanceDetailDialog, self).__init__(parent)
        self.setWindowTitle(f"Attendance Details - {date}")
        self.resize(600, 400)

        # Set up the layout
        layout = QVBoxLayout(self)

        # Add a header label
        title_label = QLabel(f"Student attendance list on {date}")
        title_label.setStyleSheet("font-size: 14pt; font-weight: bold;")
        layout.addWidget(title_label)

        # Create table
        self.tableWidget = QTableWidget()
        self.tableWidget.setColumnCount(4)
        self.tableWidget.setHorizontalHeaderLabels(["Student ID", "Full Name", "Check-in Time", "Status"])
        layout.addWidget(self.tableWidget)

        # Populate table with student attendance data
        unique_students = {}
        for record in attendance_records:
            student_id = record.get('ma_sv', '')
            if 'in' in record.get('trang_thai', '').lower() and student_id not in unique_students:
                unique_students[student_id] = {
                    'time': record.get('thoi_gian', ''),
                    'status': record.get('trang_thai', '')
                }

        self.tableWidget.setRowCount(len(unique_students))

        for row, (student_id, data) in enumerate(unique_students.items()):
            # Student ID
            self.tableWidget.setItem(row, 0, QTableWidgetItem(student_id))

            # Student name
            student_name = student_list.get(student_id, student_id)
            self.tableWidget.setItem(row, 1, QTableWidgetItem(student_name))

            # Clock-in time
            self.tableWidget.setItem(row, 2, QTableWidgetItem(data['time']))

            # Status
            self.tableWidget.setItem(row, 3, QTableWidgetItem(data['status']))

        # Resize columns to content
        self.tableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        self.tableWidget.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)

        # Add a close button
        close_button = QPushButton("Đóng")
        close_button.clicked.connect(self.accept)
        layout.addWidget(close_button)


class Ui_StatisticDialog(QDialog, Ui_StatisticDialog):
    def __init__(self, quan_ly_diem_danh):
        super(Ui_StatisticDialog, self).__init__()
        self.setupUi(self)
        self.quan_ly_diem_danh = quan_ly_diem_danh
        self.student_list = self.load_student_list()

        # Set up connections
        self.pushButtonClose.clicked.connect(self.close)
        self.pushButtonExport.clicked.connect(self.export_report)
        self.lineEditFilter.textChanged.connect(self.filter_students)
        self.checkBoxShowUnregistered.stateChanged.connect(self.filter_students)

        # Connect table cell click events
        self.tableWidgetDateStats.cellClicked.connect(self.date_stats_cell_clicked)
        self.tableWidgetStudentStats.cellClicked.connect(self.student_stats_cell_clicked)

        # Initialize tab displays
        self.tabWidget.currentChanged.connect(self.tab_changed)

        # Setup date chart
        self.date_figure = Figure(figsize=(5, 4), dpi=100)
        self.date_canvas = FigureCanvas(self.date_figure)
        layout = self.frameDateChart.layout()
        if layout is None:
            from PyQt6.QtWidgets import QVBoxLayout
            layout = QVBoxLayout(self.frameDateChart)
        layout.addWidget(self.date_canvas)

        # Setup student chart
        self.student_figure = Figure(figsize=(5, 4), dpi=100)
        self.student_canvas = FigureCanvas(self.student_figure)
        layout = self.frameStudentChart.layout()
        if layout is None:
            from PyQt6.QtWidgets import QVBoxLayout
            layout = QVBoxLayout(self.frameStudentChart)
        layout.addWidget(self.student_canvas)

        # Initial display
        self.display_date_statistics()
        self.display_student_statistics()

    def date_stats_cell_clicked(self, row, column):
        # Check if the clicked cell is in the "Số SV điểm danh" column (column index 1)
        if column == 1:
            # Get the date from the first column of this row
            date = self.tableWidgetDateStats.item(row, 0).text()

            # Get attendance records for this date
            attendance_records = self.quan_ly_diem_danh.loc_theo_ngay(date)

            # Open the detail dialog
            detail_dialog = StudentAttendanceDetailDialog(date, attendance_records, self.student_list, self)
            detail_dialog.exec()

    def student_stats_cell_clicked(self, row, column):
        # Get student ID and name from the clicked row
        student_id = self.tableWidgetStudentStats.item(row, 0).text()
        student_name = self.tableWidgetStudentStats.item(row, 1).text()

        # Create and open StudentDetailDialogExt for the selected student
        from student_detail_windowExt import Ui_StudentDetailDialogExt

        # Pass the quan_ly_diem_danh instance to provide attendance data
        detail_dialog = Ui_StudentDetailDialogExt(student_id, student_name, self.quan_ly_diem_danh)
        detail_dialog.exec()

    def tab_changed(self, index):
        if index == 0:  # Date statistics tab
            self.display_date_statistics()
        elif index == 1:  # Student statistics tab
            self.display_student_statistics()

    def display_date_statistics(self):
        # Get list of dates
        dates = self.quan_ly_diem_danh.lay_danh_sach_ngay()
        dates.sort()  # Sort chronologically

        # Get total number of students
        total_students = len(self.student_list)

        # Clear the table
        self.tableWidgetDateStats.setRowCount(0)

        date_data = []
        attendance_counts = []
        attendance_rates = []

        # Populate the table
        for i, date in enumerate(dates):
            # Get attendance for this date
            attendance = self.quan_ly_diem_danh.loc_theo_ngay(date)

            # Count unique students who checked in
            unique_students = set()
            for record in attendance:
                if 'in' in record.get('trang_thai', '').lower():
                    unique_students.add(record.get('ma_sv', ''))

            attendance_count = len(unique_students)
            attendance_rate = (attendance_count / total_students * 100) if total_students > 0 else 0

            date_data.append(date)
            attendance_counts.append(attendance_count)
            attendance_rates.append(attendance_rate)

            # Add row to table
            row_position = self.tableWidgetDateStats.rowCount()
            self.tableWidgetDateStats.insertRow(row_position)

            self.tableWidgetDateStats.setItem(row_position, 0, QTableWidgetItem(date))

            # Make the attendance count cell clickable (visual indication)
            attendance_item = QTableWidgetItem(str(attendance_count))
            attendance_item.setForeground(QBrush(QColor(0, 0, 255)))  # Blue text
            attendance_item.setToolTip("Click to view the list of checked-in students")
            self.tableWidgetDateStats.setItem(row_position, 1, attendance_item)

            self.tableWidgetDateStats.setItem(row_position, 2, QTableWidgetItem(str(total_students)))
            self.tableWidgetDateStats.setItem(row_position, 3, QTableWidgetItem(f"{attendance_rate:.2f}"))

        # Resize columns to content
        self.tableWidgetDateStats.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)

        # Create chart
        self.date_figure.clear()
        ax = self.date_figure.add_subplot(111)

        # Format date labels for x-axis
        formatted_dates = [date.split(' ')[0] if ' ' in date else date for date in date_data]

        # Bar chart
        bars = ax.bar(formatted_dates, attendance_rates, color='steelblue')

        ax.set_xlabel('Date')
        ax.set_ylabel('Attendance Rate (%)')
        ax.set_title('Attendance Rate by Date')

        # Add value labels on top of bars
        for bar in bars:
            height = bar.get_height()
            ax.text(bar.get_x() + bar.get_width() / 2., height + 1,
                    f'{height:.1f}%',
                    ha='center', va='bottom', rotation=0)

        # Rotate x-axis labels for better readability
        plt.setp(ax.get_xticklabels(), rotation=45, ha='right')

        # Tight layout to fit everything
        self.date_figure.tight_layout()

        # Refresh canvas
        self.date_canvas.draw()

    def display_student_statistics(self):
        # Get filtered student list
        filter_text = self.lineEditFilter.text().strip().lower()
        show_unregistered = self.checkBoxShowUnregistered.isChecked()

        # Get all dates with attendance
        all_dates = self.quan_ly_diem_danh.lay_danh_sach_ngay()
        total_days = len(all_dates)

        # Get all students
        all_students = self.student_list

        # Clear the table
        self.tableWidgetStudentStats.setRowCount(0)

        # Data for chart
        student_names = []
        attendance_rates = []

        # For each student, calculate attendance
        row = 0
        for student_id, student_name in all_students.items():
            # Check if student has images registered - Improved method
            has_images = self.check_student_has_registered_face(student_id)

            # Filter based on show_unregistered checkbox
            if not show_unregistered and not has_images:
                continue

            # Filter based on search text
            if filter_text and filter_text not in student_id.lower() and filter_text not in student_name.lower():
                continue

            # Count days present
            days_present = 0
            for date in all_dates:
                attendance = self.quan_ly_diem_danh.loc_theo_ngay(date)
                student_present = False

                for record in attendance:
                    if record.get('ma_sv') == student_id and 'in' in record.get('trang_thai', '').lower():
                        student_present = True
                        break

                if student_present:
                    days_present += 1

            # Calculate attendance rate
            attendance_rate = (days_present / total_days * 100) if total_days > 0 else 0

            # Add to chart data (only include top students for clarity)
            if row < 15:  # Limit to 15 students for the chart
                student_names.append(student_name if len(student_name) < 15 else student_name[:12] + '...')
                attendance_rates.append(attendance_rate)

            # Add row to table
            self.tableWidgetStudentStats.insertRow(row)

            # Make student ID and name clickable (visual indication)
            id_item = QTableWidgetItem(student_id)
            id_item.setForeground(QBrush(QColor(0, 0, 255)))  # Blue text
            id_item.setToolTip("Click to view student's attendance history")
            self.tableWidgetStudentStats.setItem(row, 0, id_item)

            name_item = QTableWidgetItem(student_name)
            name_item.setForeground(QBrush(QColor(0, 0, 255)))  # Blue text
            name_item.setToolTip("Click to view student's attendance history")
            self.tableWidgetStudentStats.setItem(row, 1, name_item)

            self.tableWidgetStudentStats.setItem(row, 2, QTableWidgetItem(str(days_present)))
            self.tableWidgetStudentStats.setItem(row, 3, QTableWidgetItem(str(total_days)))
            self.tableWidgetStudentStats.setItem(row, 4, QTableWidgetItem(f"{attendance_rate:.2f}"))

            registered_item = QTableWidgetItem("Có" if has_images else "Không")
            if not has_images:
                registered_item.setBackground(QBrush(QColor(255, 230, 230)))  # Light red background for not registered
            else:
                registered_item.setBackground(QBrush(QColor(230, 255, 230)))  # Light green background for registered

            self.tableWidgetStudentStats.setItem(row, 5, registered_item)

            row += 1

        # Resize columns to content
        self.tableWidgetStudentStats.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        self.tableWidgetStudentStats.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)

        # Create student attendance chart (if we have data)
        self.student_figure.clear()

        if student_names and attendance_rates:
            ax = self.student_figure.add_subplot(111)

            # Horizontal bar chart - better for student names
            bars = ax.barh(student_names, attendance_rates, color='lightcoral')

            ax.set_xlabel('Attendance Rate (%)')
            ax.set_ylabel('Students')
            ax.set_title('Attendance Rate by Student')

            # Add value labels on bars
            for bar in bars:
                width = bar.get_width()
                ax.text(width + 1, bar.get_y() + bar.get_height() / 2.,
                        f'{width:.1f}%',
                        ha='left', va='center')

            # Set limit to 100%
            ax.set_xlim(0, 105)

            # Tight layout
            self.student_figure.tight_layout()

        # Refresh canvas
        self.student_canvas.draw()

    def filter_students(self):
        # Re-display student statistics with the current filter
        self.display_student_statistics()

    def load_student_list(self):
        # Load student list from CSV file
        students = {}
        if os.path.exists('../dataset/students.csv'):
            try:
                with open('../dataset/students.csv', newline='', encoding='utf-8') as f:
                    reader = csv.reader(f)
                    for row in reader:
                        if len(row) >= 2:
                            student_id = row[0].strip()
                            student_name = row[1].strip() if row[1].strip() else student_id
                            students[student_id] = student_name
            except Exception as e:
                print(f"Error reading students.csv: {e}")
        return students

    def check_student_has_registered_face(self, student_id):
        """
        Improved method to check if a student has registered face images
        Now directly looks for the presence of the student's image file in the ImagesAttendance folder
        """
        # Check if the student image exists in ImagesAttendance folder
        image_path = os.path.join("../ImagesAttendance", f"{student_id}.jpg")
        if os.path.exists(image_path):
            return True

        # Also check alternate image extensions
        for ext in ['.png', '.jpeg']:
            image_path = os.path.join("../ImagesAttendance", f"{student_id}{ext}")
            if os.path.exists(image_path):
                return True

        return False

    def check_student_has_images(self, student_id):
        """Legacy method kept for compatibility"""
        return self.check_student_has_registered_face(student_id)

    def export_report(self):
        # Export statistics to Excel file
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Save Report", "", "Excel Files (*.xlsx)"
        )

        if not file_path:
            return

        if not file_path.endswith('.xlsx'):
            file_path += '.xlsx'

        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils.dataframe import dataframe_to_rows

            # Create a new workbook
            wb = Workbook()

            # Create date statistics sheet
            ws_date = wb.active
            ws_date.title = "Attendance Statistics by Date"

            # Add header
            ws_date.append(["Attendance Report by Date"])
            ws_date.append([])
            ws_date.append(["Date", "Checked-in Students", "Total Students", "Rate (%)"])

            # Add data
            for row in range(self.tableWidgetDateStats.rowCount()):
                date = self.tableWidgetDateStats.item(row, 0).text()
                attendance_count = int(self.tableWidgetDateStats.item(row, 1).text())
                total_students = int(self.tableWidgetDateStats.item(row, 2).text())
                rate = float(self.tableWidgetDateStats.item(row, 3).text())

                ws_date.append([date, attendance_count, total_students, rate])

            # Create student statistics sheet
            ws_student = wb.create_sheet("Attendance Statistics by Student")

            # Add header
            ws_student.append(["Attendance Report by Student"])
            ws_student.append([])
            ws_student.append(
                ["Student ID", "Full Name", "Check-in Sessions", "Total Sessions", "Rate (%)", "Registered Face"])

            # Add data
            for row in range(self.tableWidgetStudentStats.rowCount()):
                student_id = self.tableWidgetStudentStats.item(row, 0).text()
                student_name = self.tableWidgetStudentStats.item(row, 1).text()
                days_present = int(self.tableWidgetStudentStats.item(row, 2).text())
                total_days = int(self.tableWidgetStudentStats.item(row, 3).text())
                rate = float(self.tableWidgetStudentStats.item(row, 4).text())
                registered = self.tableWidgetStudentStats.item(row, 5).text()

                ws_student.append([student_id, student_name, days_present, total_days, rate, registered])

            # Format headers
            for ws in [ws_date, ws_student]:
                # Title formatting
                ws['A1'].font = Font(bold=True, size=14)
                ws.merge_cells('A1:D1')
                ws['A1'].alignment = Alignment(horizontal='center')

                # Column headers formatting
                header_row = 3
                for cell in ws[header_row]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="E0EEFF", end_color="E0EEFF", fill_type="solid")

                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0

                    # Kiểm tra xem có phải đối tượng MergedCell không
                    if hasattr(column[0], 'column_letter'):
                        column_letter = column[0].column_letter
                    else:
                        continue  # Bỏ qua cột này nếu không có thuộc tính column_letter

                    for cell in column:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))

                    adjusted_width = max_length + 2
                    ws.column_dimensions[column_letter].width = adjusted_width

            # Save the workbook
            wb.save(file_path)

            QMessageBox.information(self, "Success", f"Report saved successfully to {file_path}")

        except Exception as e:
            QMessageBox.warning(self, "Error", f"Error exporting report: {str(e)}")


# For testing
if __name__ == "__main__":
    import os
    import sys
    sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    from models.attendance_manager import QuanLyDiemDanh

    app = QApplication(sys.argv)
    quan_ly_diem_danh = QuanLyDiemDanh()
    quan_ly_diem_danh.doc_tu_csv('../dataset/Attendance.csv')
    dialog = Ui_StatisticDialog(quan_ly_diem_danh)
    dialog.show()
    sys.exit(app.exec())