# File: manage_windowExt.py
import os
import sys
from PyQt6.QtWidgets import (QDialog, QApplication, QPushButton, QTableWidgetItem,
                             QHeaderView, QFileDialog, QMessageBox, QTableWidget)
from PyQt6.QtGui import QColor, QBrush
from PyQt6.uic import loadUi
from attendance_manager import QuanLyDiemDanh
from statistic_windowExt import Ui_StatisticDialog
from student_detail_window import Ui_StudentDetailDialog

class Ui_ManageDialog(QDialog):
    def __init__(self):
        super(Ui_ManageDialog, self).__init__()
        loadUi("manage_window.ui", self)
        self.quan_ly_diem_danh = QuanLyDiemDanh()
        self.quan_ly_diem_danh.doc_tu_csv('Attendance.csv')
        self.pushButtonClose.clicked.connect(self.close)
        self.pushButtonImport.clicked.connect(self.import_excel)
        self.pushButtonStatistic.clicked.connect(self.open_statistic)
        self.pushButtonShowAll.clicked.connect(self.hien_thi_tat_ca_sinh_vien)
        self.thiet_lap_bang()
        self.hien_thi_danh_sach_ngay()
        self.selected_date = None
        if self.quan_ly_diem_danh.lay_danh_sach_ngay():
            self.selected_date = self.quan_ly_diem_danh.lay_danh_sach_ngay()[0]

    def thiet_lap_bang(self):
        self.tableWidget.setColumnCount(5)
        self.tableWidget.setHorizontalHeaderLabels(["Student ID", "Name", "Time", "Status", "Image registed"])
        header = self.tableWidget.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)

        # Add click handler for single-click selection
        self.tableWidget.clicked.connect(self.handle_student_click)

    def hien_thi_danh_sach_ngay(self):
        # Clear existing buttons
        for i in reversed(range(self.dateButtonLayout.count())):
            widget = self.dateButtonLayout.itemAt(i).widget()
            if widget is not None:
                widget.deleteLater()
        danh_sach_ngay = self.quan_ly_diem_danh.lay_danh_sach_ngay()
        # Sort the dates to display newest first
        danh_sach_ngay.sort(reverse=True)

        for ngay in danh_sach_ngay:
            button = QPushButton(ngay)
            button.setMinimumHeight(30)
            button.clicked.connect(lambda checked, n=ngay: self.hien_thi_diem_danh_theo_ngay(n))
            self.dateButtonLayout.addWidget(button)

        if danh_sach_ngay:
            self.hien_thi_diem_danh_theo_ngay(danh_sach_ngay[0])

    def hien_thi_tat_ca_sinh_vien(self):
        # Update label to indicate we're showing all students
        self.label_selected_date.setText("Student List")

        # Update table columns - remove attendance-related columns
        self.tableWidget.setColumnCount(3)
        self.tableWidget.setHorizontalHeaderLabels(["Student ID", "Name", "Image registed"])

        # Configure column widths
        header = self.tableWidget.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)

        # Get the student list from the manager
        danh_sach_sv = self.quan_ly_diem_danh.lay_danh_sach_sinh_vien()

        # Set row count
        self.tableWidget.setRowCount(len(danh_sach_sv))

        # Fill the table with student information
        for row, sv in enumerate(danh_sach_sv):
            ma_sv = sv.ma_sv
            ho_ten = sv.ho_ten

            # Set student ID and name
            self.tableWidget.setItem(row, 0, QTableWidgetItem(ma_sv))
            self.tableWidget.setItem(row, 1, QTableWidgetItem(ho_ten))

            # Check if student has registered a photo
            da_dang_ky_anh = sv.da_dang_ky_anh()
            item_dang_ky = QTableWidgetItem("Yes" if da_dang_ky_anh else "No")
            self.tableWidget.setItem(row, 2, item_dang_ky)

            # Highlight rows based on photo registration statusa
            if not da_dang_ky_anh:
                for col in range(3):
                    cell_item = self.tableWidget.item(row, col)
                    cell_item.setBackground(QBrush(QColor(255, 200, 200)))  # Red for no photo

        # Connect double-click signal to open student detail
        self.tableWidget.itemDoubleClicked.connect(self.open_student_detail)

    def hien_thi_diem_danh_theo_ngay(self, ngay):
        self.selected_date = ngay
        self.label_selected_date.setText(f"Date: {ngay}")

        # Reset table structure for attendance view with both check-in and check-out columns
        self.tableWidget.setColumnCount(6)
        self.tableWidget.setHorizontalHeaderLabels([
            "Student ID", "Name",
            "Check-in", "Status In",
            "Check-out", "Status Out"
        ])

        header = self.tableWidget.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(5, QHeaderView.ResizeMode.ResizeToContents)

        # First, load all students from students.csv to ensure we have correct name mappings
        student_names = {}
        if os.path.exists('students.csv'):
            try:
                with open('students.csv', newline='', encoding='utf-8') as f:
                    import csv
                    reader = csv.reader(f)
                    for row in reader:
                        if len(row) >= 2:
                            student_names[row[0]] = row[1]
            except Exception as e:
                print(f"Lỗi khi đọc file students.csv: {e}")

        # Get all attendance records for the selected date
        danh_sach_diem_danh = self.quan_ly_diem_danh.loc_theo_ngay(ngay)

        # Create a dictionary to group records by student ID
        sv_dict = {}
        for item in danh_sach_diem_danh:
            ma_sv = item['ma_sv']
            if ma_sv not in sv_dict:
                sv_dict[ma_sv] = {'check_in': None, 'check_out': None}

            # Classify the record as check-in or check-out based on status
            trang_thai = item.get('trang_thai', '').lower()
            if 'in' in trang_thai:
                sv_dict[ma_sv]['check_in'] = item
            elif 'out' in trang_thai:
                sv_dict[ma_sv]['check_out'] = item
            else:
                # If no status information, default to check-in
                if sv_dict[ma_sv]['check_in'] is None:
                    sv_dict[ma_sv]['check_in'] = item
                else:
                    sv_dict[ma_sv]['check_out'] = item

        # Make sure all students from students.csv are in the dictionary
        for ma_sv, ho_ten in student_names.items():
            if ma_sv not in sv_dict:
                sv_dict[ma_sv] = {'check_in': None, 'check_out': None}

        # Display the grouped data on the table
        self.tableWidget.setRowCount(len(sv_dict))

        row = 0
        for ma_sv, data in sv_dict.items():
            check_in_data = data['check_in']
            check_out_data = data['check_out']

            # Set student ID
            self.tableWidget.setItem(row, 0, QTableWidgetItem(ma_sv))

            # Set student name - Make sure to use the name from our mapping
            ho_ten = student_names.get(ma_sv, ma_sv)  # Default to ID if name not found
            self.tableWidget.setItem(row, 1, QTableWidgetItem(ho_ten))

            # Background color for missing data
            color_absent = QColor(255, 200, 200)  # Light red

            # Handle check-in data
            if check_in_data:
                # Extract time from full timestamp
                thoi_gian_parts = check_in_data.get('thoi_gian', '').split() if check_in_data.get('thoi_gian') else ["",
                                                                                                                     ""]
                if len(thoi_gian_parts) > 1:
                    thoi_gian_hien_thi = thoi_gian_parts[1]
                else:
                    thoi_gian_hien_thi = check_in_data.get('thoi_gian', '')

                # Set check-in time
                item_checkin_time = QTableWidgetItem(thoi_gian_hien_thi)
                self.tableWidget.setItem(row, 2, item_checkin_time)

                # Set check-in status
                trang_thai_in = check_in_data.get('trang_thai', 'Clock In')
                item_checkin_status = QTableWidgetItem(trang_thai_in)
                self.tableWidget.setItem(row, 3, item_checkin_status)
            else:
                # No check-in data
                item_checkin_time = QTableWidgetItem("Not check-in")
                item_checkin_status = QTableWidgetItem("Absence")
                self.tableWidget.setItem(row, 2, item_checkin_time)
                self.tableWidget.setItem(row, 3, item_checkin_status)
                item_checkin_time.setBackground(QBrush(color_absent))
                item_checkin_status.setBackground(QBrush(color_absent))

            # Handle check-out data
            if check_out_data:
                # Extract time from full timestamp
                thoi_gian_parts = check_out_data.get('thoi_gian', '').split() if check_out_data.get('thoi_gian') else [
                    "", ""]
                if len(thoi_gian_parts) > 1:
                    thoi_gian_hien_thi = thoi_gian_parts[1]
                else:
                    thoi_gian_hien_thi = check_out_data.get('thoi_gian', '')

                # Set check-out time
                item_checkout_time = QTableWidgetItem(thoi_gian_hien_thi)
                self.tableWidget.setItem(row, 4, item_checkout_time)

                # Set check-out status
                trang_thai_out = check_out_data.get('trang_thai', 'Clock Out')
                item_checkout_status = QTableWidgetItem(trang_thai_out)
                self.tableWidget.setItem(row, 5, item_checkout_status)
            else:
                # No check-out data
                item_checkout_time = QTableWidgetItem("Not check-out")
                item_checkout_status = QTableWidgetItem("Absence")
                self.tableWidget.setItem(row, 4, item_checkout_time)
                self.tableWidget.setItem(row, 5, item_checkout_status)
                item_checkout_time.setBackground(QBrush(color_absent))
                item_checkout_status.setBackground(QBrush(color_absent))

            row += 1

    def import_excel(self):
        file_dialog = QFileDialog()
        file_path, _ = file_dialog.getOpenFileName(
            self, "Chọn file Excel", "", "Excel Files (*.xlsx *.xls)"
        )
        if file_path:
            try:
                # Sử dụng phương thức từ module readExcel
                from openpyxl import load_workbook

                # Đọc file Excel
                wb = load_workbook(filename=file_path, read_only=True)
                ws = wb.active  # Lấy sheet đầu tiên

                # Xác định header row
                header_row = next(ws.iter_rows(values_only=True))
                header_row = [str(cell).lower() if cell else "" for cell in header_row]

                # Tìm các cột chứa StudentID và FullName
                student_id_col = None
                fullname_col = None

                for idx, header in enumerate(header_row):
                    if header in ["studentid", "student_id", "masv", "mã sv", "ma_sv", "id"]:
                        student_id_col = idx
                    elif header in ["fullname", "full_name", "hoten", "họ tên", "ho_ten", "name"]:
                        fullname_col = idx

                # Kiểm tra nếu không tìm thấy các cột cần thiết
                if student_id_col is None or fullname_col is None:
                    QMessageBox.warning(self, "Lỗi", "Không tìm thấy cột StudentID hoặc FullName trong file Excel.")
                    return False

                # Đọc dữ liệu từ file Excel
                student_list = []
                for row in list(ws.iter_rows(values_only=True))[1:]:  # Bỏ qua header
                    if row[student_id_col] and row[fullname_col]:  # Kiểm tra dữ liệu hợp lệ
                        student_id = str(row[student_id_col]).strip()
                        full_name = str(row[fullname_col]).strip()
                        student_list.append([student_id, full_name])

                # Lưu vào file students.csv
                if student_list:
                    with open('students.csv', 'w', newline='', encoding='utf-8') as f:
                        import csv
                        writer = csv.writer(f)
                        for student in student_list:
                            writer.writerow(student)

                    # Cập nhật danh sách sinh viên
                    self.quan_ly_diem_danh.load_student_info()

                    QMessageBox.information(self, "Thành công", f"Đã nhập {len(student_list)} sinh viên từ file Excel.")
                    self.hien_thi_tat_ca_sinh_vien()
                    return True
                else:
                    QMessageBox.warning(self, "Cảnh báo", "Không tìm thấy dữ liệu sinh viên trong file Excel.")
                    return False

            except Exception as e:
                QMessageBox.warning(self, "Lỗi", f"Không thể xử lý file Excel: {str(e)}")
                return False
    def open_statistic(self):
        statistic_dialog = Ui_StatisticDialog(self.quan_ly_diem_danh)
        statistic_dialog.exec()

    def handle_student_click(self, index):
        # Get the row of the clicked item
        row = index.row()

        # Get student ID and name from the table
        student_id = self.tableWidget.item(row, 0).text()
        student_name = self.tableWidget.item(row, 1).text()

        # Create and show the student detail dialog
        from student_detail_windowExt import Ui_StudentDetailDialogExt
        detail_dialog = Ui_StudentDetailDialogExt(student_id, student_name, self.quan_ly_diem_danh)
        detail_dialog.exec()

        # Refresh the display after dialog closes
        if self.selected_date:
            self.hien_thi_diem_danh_theo_ngay(self.selected_date)
        else:
            self.hien_thi_tat_ca_sinh_vien()
    def open_student_detail(self, item):
        # Get the row of the clicked item
        row = item.row()

        # Get student ID from the first column
        student_id = self.tableWidget.item(row, 0).text()

        # Get student name from the second column
        student_name = self.tableWidget.item(row, 1).text()

        # Import the student detail window here to avoid circular imports
        from student_detail_windowExt import Ui_StudentDetailDialogExt

        # Create and show the student detail dialog
        detail_dialog = Ui_StudentDetailDialogExt(student_id, student_name, self.quan_ly_diem_danh)
        detail_dialog.exec()

        # Refresh the student list after closing the detail dialog
        if self.selected_date:
            self.hien_thi_diem_danh_theo_ngay(self.selected_date)
        else:
            self.hien_thi_tat_ca_sinh_vien()

def main():
    app = QApplication(sys.argv)
    window = Ui_ManageDialog()
    window.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()